#!/usr/bin/env python3

__author__ = 'Jeff'
__copyright__ = 'Copyright 2023, Kather Lab'
__license__ = 'MIT'
__version__ = '0.1.0'
__maintainer__ = ['Jeff']
__email__ = 'jiefu.zhu@tu-dresden.de'

import h5py
import numpy as np
import openpyxl
from pathlib import Path

import pandas as pd

from cluster_and_plot import DataClustering

h5outpath = '/mnt/sda1/feature-visualizations/Xiyue-Wang/'
# inlude all the files in the folder
import os
import glob
file_list = glob.glob('/mnt/sda1/feature-visualizations/Xiyue-Wang/*.h5')


def construct_label_dict(num=2):
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(
        "/mnt/sda1/feature-visualizations/JULIEN-POLE-Hypermutated Solid Tumors 2022/PARIS_full_CLINI.xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook.active
    # put first row and second row into a dictionary
    dict = {}
    # Iterate the loop to read the cell values
    for row in range(2, worksheet.max_row):
        dict[worksheet.cell(row, 1).value] = worksheet.cell(row, num).value
    #print(dict)
    return dict
def get_feature_label(patient_name, dict):
    # read xlsx file and get the corresponding label to patient_name

    # return the label
    return dict[patient_name]




def construct_dataframe(colume = 2):
    X = np.zeros((0, 2048))
    y = np.zeros((0, 1))

    for file in file_list[:2]:
        #print(file)
        f = h5py.File(file, 'r')
        #print(list(f.keys()))
        patient_name = Path(file).stem.split('.')[-1]
        #print(patient_name)
        label = get_feature_label(patient_name, construct_label_dict(colume))
        # generate a label array with the same length as the feature array
        #print(label)

        ds_features = f.get('feats')
        #print(ds_features.shape)
        label_array = np.full((ds_features.shape[0], 1), label)
        #print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    #print(X.shape)
    print(y)
    #print(y.shape)
    return X, y

def construct_dataframe_perpatient_mean():
    X = np.zeros((0, 2048))
    y = np.zeros((0, 1))

    for file in file_list:
        print(file)
        f = h5py.File(file, 'r')
        print(list(f.keys()))
        patient_name = Path(file).stem.split('.')[-1]
        print(patient_name)
        label = get_feature_label(patient_name, construct_label_dict())

        print(label)
        ds_features = f.get('feats')
        # get mean of ds_features
        ds_features = np.mean(ds_features, axis=0)

        # generate a label array with the same length as the feature array
        label_array = np.full((1, 1), label)

        print(ds_features.shape)
        print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    print(X.shape)
    print(X)
    print(y.shape)
    return X, y

X, y = construct_dataframe(colume=2)
#X, y = construct_dataframe_perpatient_mean()
dc = DataClustering('out_dir', X, y)

#dc.run_tSNE()
dc.run_pca()
dc.plot_scatter()
#dc.plot_imgs_withinput()