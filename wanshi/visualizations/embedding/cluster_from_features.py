#!/usr/bin/env python3

__author__ = 'Jeff'
__copyright__ = 'Copyright 2023, Kather Lab'
__license__ = 'MIT'
__version__ = '0.1.0'
__maintainer__ = ['Jeff']
__email__ = 'jiefu.zhu@tu-dresden.de'

import argparse

import h5py
import numpy as np
import openpyxl
from pathlib import Path

import pandas as pd

from cluster_and_plot import DataClustering



def construct_label_dict(num):
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(
        "/mnt/sda1/vis_embedding_data/JULIEN_POLE_2022_CLINI_P806046_07.02.23.xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook.active
    # put first row and second row into a dictionary
    dict = {}
    # Iterate the loop to read the cell values
    for row in range(2, worksheet.max_row):
        dict[worksheet.cell(row, 1).value] = worksheet.cell(row, num).value
    #print(dict)
    return dict
def get_feature_label(patient_name, dict):
    # read xlsx file and get the corresponding label to patient_name

    # return the label
    return dict[patient_name]




def construct_dataframe(file_list, colume):
    X = np.zeros((0, 2048))
    y = np.zeros((0, 1))

    for file in file_list:
        #print(file)
        f = h5py.File(file, 'r')
        #print(list(f.keys()))
        patient_name = Path(file).stem.split('.')[-1]
        #print(patient_name)
        try:
            label = get_feature_label(patient_name, construct_label_dict(colume))
        except KeyError:
            continue
        # generate a label array with the same length as the feature array
        print(label)

        ds_features = f.get('feats')
        #print(ds_features.shape)
        label_array = np.full((ds_features.shape[0], 1), label)
        #print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    #print(X.shape)
    print(y)
    #print(y.shape)
    return X, y

def construct_dataframe_compare_feature_ext(file_list, dimension):
    X = np.zeros((0, dimension))
    y = np.zeros((0, 1))
    shape_list = []
    for file in file_list[:50]:
        #print(file)
        f = h5py.File(file, 'r')
        # get the parent folder name of the file
        parent_folder = Path(file).parent.name
        label = str(parent_folder)

        # generate a label array with the same length as the feature array
        #print(label)

        ds_features = f.get('feats')
        #randomly select 1000 features
        #print(ds_features.shape)
        shape_list.append(ds_features.shape[0])
        label_array = np.full((ds_features.shape[0], 1), label)
        #print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    print(X.shape)
    print(y)
    #print(y.shape)
    # get the minimum shape and max shape from the shape_list
    min_shape = min(shape_list)
    max_shape = max(shape_list)
    print('####################')
    print(min_shape)
    print(max_shape)
    return X, y

def construct_dataframe_compare_feature_ext_flatten(file_list, dimension, sample_num):
    X = np.zeros((0, dimension))
    y = np.zeros((0, 1))
    shape_list = []
    for file in file_list[:500]:
        #print(file)
        f = h5py.File(file, 'r')
        # get the parent folder name of the file
        parent_folder = Path(file).parent.name
        label = str(parent_folder)

        # generate a label array with the same length as the feature array
        #print(label)

        ds_features = f.get('feats')
        #randomly sample 740 features
        ds_features = ds_features[:sample_num, :]
        #ds_features = ds_features[np.random.choice(ds_features.shape[0], 740, replace=False), :]
        #flatten the features
        ds_features = ds_features.flatten()
        print('ds_features.shape',ds_features.shape)
        shape_list.append(ds_features.shape[0])
        label_array = np.full((1, 1), label)
        #print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    #print('X.shape',X.shape)
    #print('y.shape',y.shape)
    #print(y)
    #print(y.shape)
    # get the minimum shape and max shape from the shape_list
    #min_shape = min(shape_list)
    #max_shape = max(shape_list)
    #print('####################')
    #print(min_shape)
    #print(max_shape)
    print('####################')
    min_shape = min(shape_list)
    max_shape = max(shape_list)
    print(min_shape)
    print(max_shape)
    return X, y
def construct_dataframe_perpatient_mean():
    X = np.zeros((0, 2048))
    y = np.zeros((0, 1))

    for file in file_list:
        print(file)
        f = h5py.File(file, 'r')
        print(list(f.keys()))
        patient_name = Path(file).stem.split('.')[-1]
        print(patient_name)
        label = get_feature_label(patient_name, construct_label_dict())
        #label = 1
        print(label)
        ds_features = f.get('feats')
        # get mean of ds_features
        ds_features = np.mean(ds_features, axis=0)

        # generate a label array with the same length as the feature array
        label_array = np.full((1, 1), label)

        print(ds_features.shape)
        print(label_array.shape)

        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()

    #print(X.shape)
    #print(X)
    #print(y.shape)
    return X, y

def get_dimension(file):
    f = h5py.File(file, 'r')
    ds_features = f.get('feats')
    return ds_features.shape[1]

if __name__ == '__main__':
    for cohort in ['dachs']:
        # parse arguments
        parser = argparse.ArgumentParser()
        #parser.add_argument('--feature_dir', type=str, default='/mnt/sda1/vis_embedding_data/paris/',
        #                    help='path to the h5 file')

        #args = parser.parse_args()
        #data_dir = Path(args.feature_dir)
        data_dir1 = Path('/mnt/sda1/vis_embedding_data/{}/RetCCL'.format(cohort))
        data_dir2 = Path('/mnt/sda1/vis_embedding_data/{}/CTrans'.format(cohort))

        # get all the files in the data_dir iteratively

        h5_file_list1 = [f for f in data_dir1.iterdir() if f.is_file() and f.suffix == '.h5']
        h5_file_list2 = [f for f in data_dir2.iterdir() if f.is_file() and f.suffix == '.h5']
        #h5_file_list1.extend(h5_file_list2)
        #print(h5_file_list1)    #X, y = construct_dataframe(h5_file_list, colume=7)
        #X, y = construct_dataframe_compare_feature_ext(h5_file_list1, dimension=get_dimension(h5_file_list1[0]))
        #X2, y2 = construct_dataframe_compare_feature_ext(h5_file_list2, dimension=get_dimension(h5_file_list2[0]))
        sample_num = 500
        X, y = construct_dataframe_compare_feature_ext_flatten(h5_file_list1, dimension=sample_num*2048, sample_num=sample_num)
        print('123')
        X2, y2 = construct_dataframe_compare_feature_ext_flatten(h5_file_list2, dimension=sample_num*768, sample_num=sample_num)

        #X = np.concatenate((X, X2), axis=1)
        y_final = np.concatenate((y, y2), axis=0)
        print(X.shape)
        print(y.shape)
        # X, y = construct_dataframe_perpatient_mean()
        dc1 = DataClustering('out_dir', X, y)

        #dc.run_tSNE()
        #dc.plot_scatter('tSNE_')

        x2d1 = dc1.run_pca()
        #print(x2d1.shape)
        dc2 = DataClustering('out_dir', X2, y2)
        x2d2 = dc2.run_pca()
        #print(x2d2.shape)
        #X_pca = np.hstack((x2d1, x2d2))
        X_pca = np.concatenate((x2d1, x2d2), axis=0)

        print(X_pca.shape)
        # random sample 1000 data points
        #X_pca = X_pca[np.random.choice(X_pca.shape[0], 100000, replace=False), :]
        print(y_final.shape)
        #y_final = y_final[np.random.choice(y_final.shape[0], 100000, replace=False), :]
        dc1.get_feature_label(X_pca, y_final)
        dc1.plot_scatter('PCA_', "dachs feature clustering flatten", size=10)
        # dc.plot_imgs_withinput()
        #################
        '''
        dc1 = DataClustering('out_dir', X, y)

        x2d1 = dc1.run_tSNE()
        print(x2d1.shape)
        dc2 = DataClustering('out_dir', X2, y2)
        x2d2 = dc2.run_tSNE()
        print(x2d2.shape)
        #X_pca = np.hstack((x2d1, x2d2))
        X_tSNE = np.concatenate((x2d1, x2d2), axis=0)
        #X_tSNE = X_tSNE[np.random.choice(X_tSNE.shape[0], 10000, replace=False), :]
        #y_final = y_final[np.random.choice(y_final.shape[0], 10000, replace=False), :]
        print(X_tSNE.shape)
        print(y_final.shape)
        dc1.get_feature_label(X_tSNE, y_final)
        dc1.plot_scatter('tSNE_', "{} feature clustering".format(cohort), "{} feature clustering".format(cohort))
        '''