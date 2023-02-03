#!/usr/bin/env python3

__author__ = 'Jeff'
__copyright__ = 'Copyright 2023, Kather Lab'
__license__ = 'MIT'
__version__ = '0.1.0'
__maintainer__ = ['Jeff']
__email__ = 'jiefu.zhu@tu-dresden.de'

import openpyxl

def read_xlsx(file_path, column_num):
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(file_path)
    # Define variable to read the active sheet:
    worksheet = wookbook.active
    # put column into a list
    list = []
    # Iterate the loop to read the cell values
    for row in range(2, worksheet.max_row):
        list.append(worksheet.cell(row, column_num).value)
    return list


def get_label_list_for_each_series(list):
    label_list = []
    for file_name in list:
        label = file_name.split('/')[-2]
        label_list.append(label)
    return label_list


def get_label_list_for_each_series_folder(list):
    label_list = []
    for file_name in list:
        print(file_name)
        try:
            label = file_name.split('/')[-2].split('-')[-2]
        except IndexError:
            label = file_name.split('//')[-2].split('-')[-2]
        label_list.append(label)
    return label_list


def calculate_label_distribution(label_list):
    other_count = 0
    label_set = set(label_list)
    label_distribution = {}
    for label in label_set:
        label_distribution[label] = label_list.count(label)
    return label_distribution


def read_lable_distribution_and_merge_key_with_value_less_then_th_to_other(label_distribution, treshold):
    other_count = 0
    label_distribution_copy = label_distribution.copy()
    for key in label_distribution_copy:
        if label_distribution[key] < treshold:
            other_count += label_distribution[key]
            del label_distribution[key]
    label_distribution['other'] = other_count
    return label_distribution


def plot_label_distribution(label_distribution, name, rotation=0):
    import matplotlib.pyplot as plt
    # set the size of the plot
    plt.figure(figsize=(10, 5))
    barlist = plt.bar(label_distribution.keys(), label_distribution.values(), color='g')
    barlist[0].set_color('r')

    # set the title of the plot as "series distribution"
    plt.title(name)
    # set the x axis label as "series name"
    plt.xlabel("series name")
    plt.xticks(rotation=rotation)

    # set the y axis label as "number of DICOM files"
    plt.ylabel("number of DICOM files")
    plt.show()
    plt.gca().get_xticklabels()[0].set_color('red')
    # save the plot as a png file
    plt.savefig(name + '.png')


if __name__ == '__main__':
    list = read_xlsx(
        "/mnt/sda1/swarm-learning/radiology-dataset/original-dataset/Breast-Cancer-MRI-filepath_filename-mapping.xlsx",
        4)
    label_list = get_label_list_for_each_series_folder(list)
    label_distribution = calculate_label_distribution(label_list)
    # label_distribution = { 'pre': 157198, 'T1': 74049, 'post_1': 157198, 'post_2': 157198, 'post_3': 155526,'post_4': 71956}
    label_distribution = read_lable_distribution_and_merge_key_with_value_less_then_th_to_other(label_distribution,
                                                                                                49000)
    # label_distribution = {'ax dyn pre': 49476,'Ax Vibrant MultiPhase': 64720, 'ax 3d dyn MP': 147610, 'ax 3d dyn': 147702, 'ax dyn 2nd pass': 49416,  'ax dynamic': 75520, 'ax dyn 1st pass': 49120, 'other': 189561}
    # {'Ax 3D T1 NON FS': 12944, 'ax t1 pre': 361, 'ax dyn 4th pass': 12792, 'ax 3d pre': 964, 'ax 3d dyn pre': 4202, 'ax t1 tse +c': 745, 'ax 3d dyn 4th pass': 4202, 'Ax Vibrant MultiPhase': 64720, 't1fl3dtradynVIEWSspair2nd pass': 480, 'ax 3d dyn MP': 147610, '1st ax dyn': 964, 'ax 3d dyn 2nd pass': 4202, 't1fl3dtradynVIEWSspair pre': 480, 'ax dyn 1st pas': 120, 't1fl3dtradynVIEWSspair3rd pass': 480, 'ax dyn pre': 49476, 'ax 3d dyn 3rd pass': 4202, 'AX IDEAL Breast': 305, '4th ax dyn': 868, 'ax t1': 44268, 'ax t1 3mm': 891, 't1fl3dtradynVIEWSspair 1st pass': 480, '2nd ax dyn': 964, 'ax 3d t1 bilateral': 12192, 'ax 3d dyn 1st pass': 4202, '3rd ax dyn': 964, 'ax t1 2mm': 268, 'ax t1 tse': 1096, 'ax 3d dyn': 147702, 'ax t1 repeat': 162, 'ax dyn 1st pass': 49120, 'ax dynamic': 75520, 'ax t1 +c': 817, 't1fl3dtradynVIEWSspair4th pass': 480, 'ax dyn': 26362, 'ax dyn 2nd pass': 49416, 'ax dyn 3rd pass': 48104}
    plot_label_distribution(label_distribution, 'distribution of folder names', 13)
    plot_label_distribution(label_distribution, 'series distribution')


# pytest for each function
def test_read_xlsx():
    list = read_xlsx(
        "/mnt/sda1/swarm-learning/radiology-dataset/original-dataset/Breast-Cancer-MRI-filepath_filename-mapping.xlsx",
        4)
    assert len(list) == 157198
    assert list[0] == '/mnt/sda1/swarm-learning/radiology-dataset/original-dataset/ax dyn pre/000000.dcm'
    assert list[-1] == '/mnt/sda1/swarm-learning/radiology-dataset/original-dataset/ax dyn pre/157197.dcm'


# pylint for each function
def test_get_label_list_for_each_series_folder():
    list = read_xlsx(
        "/mnt/sda1/swarm-learning/radiology-dataset/original-dataset/Breast-Cancer-MRI-filepath_filename-mapping.xlsx",
        4)
    label_list = get_label_list_for_each_series_folder(list)
    assert len(label_list) == 157198
    assert label_list[0] == 'ax dyn pre'
    assert label_list[-1] == 'ax dyn pre'

'''
post_1 - ax dyn 1st pass / ax 3d dyn
post_2 - ax dyn 2nd pass / ax 3d dyn
post_3 - ax dyn 3rd pass / ax 3d dyn
post_4 - ax dyn 4th pass
t1 - ax t1 tse
pre - ax 3d dyn
'''