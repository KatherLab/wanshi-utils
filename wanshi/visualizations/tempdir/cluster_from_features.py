#!/usr/bin/env python3
import argparse
import h5py
import numpy as np
import openpyxl
from pathlib import Path
from cluster_and_plot import DataClustering

def construct_label_dict(num):
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(
        "/mnt/sda1/vis_embedding_data/JULIEN_POLE_2022_CLINI_P806046_07.02.23.xlsx")
    worksheet = wookbook.active
    dict = {}
    for row in range(2, worksheet.max_row):
        dict[worksheet.cell(row, 1).value] = worksheet.cell(row, num).value
    return dict

def get_feature_label(patient_name, dict):
    return dict[patient_name]

def construct_dataframe(file_list, colume):
    X = np.zeros((0, 2048))
    y = np.zeros((0, 1))

    for file in file_list:
        f = h5py.File(file, 'r')
        patient_name = Path(file).stem.split('.')[-1]
        try:
            label = get_feature_label(patient_name, construct_label_dict(colume))
        except KeyError:
            continue
        print(label)
        ds_features = f.get('feats')
        label_array = np.full((ds_features.shape[0], 1), label)
        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()
    print(y)
    return X, y

def construct_dataframe_compare_feature_ext_flatten(file_list, dimension, sample_num):
    X = np.zeros((0, dimension))
    y = np.zeros((0, 1))
    shape_list = []
    for file in file_list[:50]:#TODO: remove this limit
        #print(file)
        f = h5py.File(file, 'r')
        # get the parent folder name of the file
        parent_folder = Path(file).parent.name
        label = str(parent_folder)
        ds_features = f.get('feats')
        ds_features = ds_features[:sample_num, :]
        ds_features = ds_features.flatten()
        shape_list.append(ds_features.shape[0])
        label_array = np.full((1, 1), label)
        X = np.vstack([X, ds_features])
        y = np.vstack([y, label_array])
        f.close()
    return X, y

def get_dimension(file):
    f = h5py.File(file, 'r')
    ds_features = f.get('feats')
    return ds_features.shape[1]

if __name__ == '__main__':
    for cohort in ['dachs']:#['dachs' , 'paris', 'tcga']
        parser = argparse.ArgumentParser()
        data_dir_RetCCL = Path('/mnt/sda1/vis_embedding_data/{}/RetCCL'.format(cohort))
        data_dir_CTrans = Path('/mnt/sda1/vis_embedding_data/{}/CTrans'.format(cohort))

        sample_num = 500
        RetCCL_feature_dimention = 2048
        CTrans_feature_dimention = 768
        X, y = construct_dataframe_compare_feature_ext_flatten([f for f in data_dir_RetCCL.iterdir() if f.is_file() and f.suffix == '.h5'], dimension=sample_num*RetCCL_feature_dimention, sample_num=sample_num)
        X2, y2 = construct_dataframe_compare_feature_ext_flatten([f for f in data_dir_CTrans.iterdir() if f.is_file() and f.suffix == '.h5'], dimension=sample_num*CTrans_feature_dimention, sample_num=sample_num)
        y_final = np.concatenate((y, y2), axis=0)
        dc = DataClustering('out_dir', X, y)
        x2d1 = dc.run_pca()
        dc2 = DataClustering('out_dir', X2, y2)
        x2d2 = dc.run_pca()
        X_pca = np.concatenate((x2d1, x2d2), axis=0)
        dc.get_feature_label(X_pca, y_final)
        dc.plot_scatter('PCA_', "dachs feature clustering flatten", size=10)
